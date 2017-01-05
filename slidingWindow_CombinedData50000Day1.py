# -*- coding: utf-8 -*-
"""
Created on Fri Dec 30 23:58:03 2016

@author: kashu
"""

import pandas as pd
from sklearn import linear_model
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.backends.backend_pdf as bpdf
from sklearn.metrics import mean_squared_error
from math import sqrt
import openpyxl as opxl
#from LinearAndMultiLinearRegression import combineDay1Day2TestDay3
#from LinearAndMultiLinearRegression import combineDay2Day3TestDay1
#from LinearAndMultiLinearRegression import combineDay1Day3TestDay2


dataCompete = pd.ExcelFile('C:/MASTERS/Project/slidingWindow/combinedData50000.xlsx') 
wsday1 =  dataCompete.parse('Sheet1')[0:50006]
#wsday2 =  dataCompete.parse('Sheet2')[0:43168]
#wsday3 =  dataCompete.parse('Sheet3')[0:43037]
wb = opxl.Workbook()
pdfRegressionParametrs = bpdf.PdfPages("Analysis Plots - RegressionParametersCombinedData50000Day1.pdf")
for x in range(1,9):
    independentDay1 = np.array(wsday1[["input_heat_energy"+str(x),"input_wall_temp"+str(x),"input_mat_temp"+str(x)]])
    dependentDay1 = np.array(wsday1["output_section_temp"+str(x)])
  
    relation = "output_section_temp "+str(x) + " ~ input_heat_energy"+str(x)," input_wall_temp"+str(x)," input_mat_temp"+str(x)
   

    sheet = wb.create_sheet(title=str("output_section_temp "+str(x)))
    
    sheet.cell(row=1, column=1).value = "RMS"
    sheet.cell(row=1, column=2).value = "R square"
    sheet.cell(row=1, column=3).value = "Adj R square"
    sheet.cell(row=1, column=4).value = "Predicted R square"
    
    c=0
    interceptArrray=np.empty([47],dtype='float64')
    coeff0Array=np.empty([47],dtype='float64')
    coeff1Array=np.empty([47],dtype='float64')
    coeff2Array=np.empty([47],dtype='float64')
    RMSArray =np.empty([47],dtype='float64')
    for y in range(0,41907,900):
        
        ##Initialize the model
        LinearRegressor = linear_model.LinearRegression()
        LinearRegressor.fit(independentDay1[(0+y):(7200+y)],dependentDay1[(0+y):(7200+y)])
        predictionSeenData = LinearRegressor.predict(independentDay1[(0+y):(7200+y)])
        r2score = r2_score(dependentDay1[(0+y):(7200+y)],predictionSeenData)
        predictionNewData= LinearRegressor.predict(independentDay1[(7200+y):(8100+y)])
        
        RMS = sqrt(mean_squared_error(dependentDay1[(7200+y):(8100+y)], predictionNewData))
        predictionR2score = r2_score(dependentDay1[(7200+(y)):(8100+(y))], predictionNewData)
        adjRSquare = 1-(((1-r2score)*(7200-1))/(7200-3-1))
        
        interceptArrray[c] =  LinearRegressor.intercept_
        coeff0Array[c] = LinearRegressor.coef_[0]
        coeff1Array[c] = LinearRegressor.coef_[1]
        coeff2Array[c] = LinearRegressor.coef_[2]
        RMSArray[c] = RMS
        c=c+1
        sheet.cell(row=c+1, column=1).value = RMS
        sheet.cell(row=c+1, column=2).value = r2score
        sheet.cell(row=c+1, column=3).value = adjRSquare
        sheet.cell(row=c+1, column=4).value = predictionR2score

    plt.plot(interceptArrray, 'b*', label = 'Intercept')
    plt.plot(coeff0Array, 'ro', label = 'Coeff 0')
    plt.plot(coeff1Array, 'g^', label = 'Coeff 1')
    plt.plot(coeff2Array, 'yd', label = 'Coeff 2')
    #plt.xlabel('COunt')
    
    #graphTitle = outputTemp+str(y):   #'Zone'+str(x)+':  ('+ outputTemp + ' = ' + str(round(regSection.params[0],3)) +' + '+ str(round(regSection.params[1],3))+ '* inputEnerygy'+str(x)+ ' + '+ str(round(regSection.params[2],3)) +' * inputWallTemp'+str(x)+')'
    plt.title(relation,fontsize=7)
    plt.grid(True)
    plt.ylabel('Regression Parameters ')
    plt.legend(loc='best',fontsize=7)            
    
    #    #####Save the PDF with Graphs    
    plt.savefig(pdfRegressionParametrs, format='pdf')
    plt.close()
wb.save('Analysis Plots - SlidingWindowParametersCombinedData50000Day1.xlsx')
pdfRegressionParametrs.close()    
    